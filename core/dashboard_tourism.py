import os
import sqlite3
import threading
import time
from datetime import date, datetime

from openpyxl import load_workbook


REQUIRED_HEADERS = (
    'tourist_id', 'user_nickname', 'age', 'gender', 'attraction_name',
    'attraction_content', 'attraction_type', 'visit_date', 'stay_duration',
    'ticket_cost', 'food_cost', 'shopping_cost', 'transport_cost',
    'entertainment_cost', 'total_cost', 'group_size', 'satisfaction',
)
IMPORT_LOCK = threading.Lock()


def ensure_tourism_schema(db_path):
    os.makedirs(os.path.dirname(db_path), exist_ok=True)
    conn = sqlite3.connect(db_path)
    try:
        conn.execute('''CREATE TABLE IF NOT EXISTS tourism_visit (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            tourist_id TEXT, user_nickname TEXT, age INTEGER, gender TEXT,
            attraction_name TEXT, attraction_type TEXT, visit_date TEXT,
            stay_duration REAL, ticket_cost REAL, food_cost REAL,
            shopping_cost REAL, transport_cost REAL, entertainment_cost REAL,
            total_cost REAL, group_size INTEGER, satisfaction INTEGER
        )''')
        conn.execute('''CREATE TABLE IF NOT EXISTS tourism_import_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            file_path TEXT, file_mtime REAL, row_count INTEGER, status TEXT,
            error TEXT, imported_at INTEGER, min_date TEXT, max_date TEXT
        )''')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_tourism_visit_date ON tourism_visit(visit_date)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_tourism_type ON tourism_visit(attraction_type)')
        conn.execute('CREATE INDEX IF NOT EXISTS idx_tourism_name ON tourism_visit(attraction_name)')
        conn.commit()
    finally:
        conn.close()


def normalize_excel_text(value):
    if value is None:
        return ''
    text = str(value).strip()
    try:
        fixed = text.encode('latin1').decode('gbk')
        if fixed.count('\ufffd') < text.count('\ufffd'):
            return fixed.strip()
    except Exception:
        pass
    return text


def normalize_date(value):
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value or '').strip()
    return text[:10]


def _float(value):
    try:
        return float(value or 0)
    except (TypeError, ValueError):
        return 0.0


def _int(value):
    try:
        return int(float(value or 0))
    except (TypeError, ValueError):
        return 0


def _latest_success(conn):
    row = conn.execute(
        '''SELECT file_path, file_mtime, row_count, imported_at, min_date, max_date
        FROM tourism_import_log WHERE status = 'success' ORDER BY id DESC LIMIT 1'''
    ).fetchone()
    if not row:
        return None
    return {
        'file_path': row[0], 'file_mtime': row[1], 'row_count': row[2], 'record_count': row[2],
        'imported_at': row[3], 'import_status': 'success',
        'date_range': {'start': row[4], 'end': row[5]},
    }


def import_tourism_excel(db_path, excel_path, force=False):
    with IMPORT_LOCK:
        ensure_tourism_schema(db_path)
        if not os.path.exists(excel_path):
            return _missing_import_result(excel_path)
        file_mtime = os.path.getmtime(excel_path)
        conn = sqlite3.connect(db_path)
        try:
            latest = _latest_success(conn)
            if latest and not force and latest['file_mtime'] == file_mtime:
                return {**latest, 'success': True, 'skipped': True}
            result = _rebuild_tourism_db(conn, excel_path, file_mtime)
            conn.commit()
            return result
        except Exception as exc:
            conn.rollback()
            _record_import(conn, excel_path, file_mtime, 0, 'failed', str(exc), '', '')
            conn.commit()
            raise
        finally:
            conn.close()


def _missing_import_result(excel_path):
    return {
        'success': False, 'skipped': False, 'row_count': 0, 'file_path': excel_path,
        'error': 'tourism Excel not found', 'import_status': 'missing',
        'date_range': {'start': '', 'end': ''},
    }


def _rebuild_tourism_db(conn, excel_path, file_mtime):
    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        headers = tuple(cell for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True)))
        _validate_headers(headers)
        rows, min_date, max_date = _read_rows(sheet)
    finally:
        workbook.close()
    conn.execute('DELETE FROM tourism_visit')
    conn.executemany(_insert_sql(), rows)
    _record_import(conn, excel_path, file_mtime, len(rows), 'success', '', min_date, max_date)
    return {
        'success': True, 'skipped': False, 'row_count': len(rows),
        'record_count': len(rows),
        'file_path': excel_path, 'file_mtime': file_mtime, 'import_status': 'success',
        'date_range': {'start': min_date, 'end': max_date},
    }


def _validate_headers(headers):
    missing = [name for name in REQUIRED_HEADERS if name not in headers]
    if missing:
        raise ValueError(f'missing Excel columns: {", ".join(missing)}')


def _read_rows(sheet):
    rows = []
    min_date = ''
    max_date = ''
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row or not row[0]:
            continue
        item = _row_to_record(row)
        visit_date = item[6]
        min_date = visit_date if not min_date or visit_date < min_date else min_date
        max_date = visit_date if not max_date or visit_date > max_date else max_date
        rows.append(item)
    return rows, min_date, max_date


def _row_to_record(row):
    return (
        normalize_excel_text(row[0]), normalize_excel_text(row[1]), _int(row[2]),
        normalize_excel_text(row[3]), normalize_excel_text(row[4]),
        normalize_excel_text(row[6]), normalize_date(row[7]), _float(row[8]),
        _float(row[9]), _float(row[10]), _float(row[11]), _float(row[12]),
        _float(row[13]), _float(row[14]), _int(row[15]), _int(row[16]),
    )


def _insert_sql():
    return '''INSERT INTO tourism_visit (
        tourist_id, user_nickname, age, gender, attraction_name, attraction_type,
        visit_date, stay_duration, ticket_cost, food_cost, shopping_cost,
        transport_cost, entertainment_cost, total_cost, group_size, satisfaction
    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)'''


def _record_import(conn, file_path, file_mtime, row_count, status, error, min_date, max_date):
    conn.execute(
        '''INSERT INTO tourism_import_log
        (file_path, file_mtime, row_count, status, error, imported_at, min_date, max_date)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)''',
        (file_path, file_mtime, row_count, status, error, int(time.time()), min_date, max_date),
    )


def latest_source(db_path):
    ensure_tourism_schema(db_path)
    conn = sqlite3.connect(db_path)
    try:
        latest = _latest_success(conn)
        return latest or {
            'row_count': 0, 'record_count': 0, 'import_status': 'not_imported',
            'date_range': {'start': '', 'end': ''},
        }
    finally:
        conn.close()
