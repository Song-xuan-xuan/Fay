import csv
import io

from flask import jsonify, request

from core import auth_service
from core.tourism_recommendation_service import TourismRecommendationService


def json_data():
    return request.get_json(silent=True) or {}


def service():
    return TourismRecommendationService()


def current_user():
    return auth_service.current_user()


def login_required_response():
    if not current_user():
        return jsonify({'error': '推荐功能需要登录'}), 401
    return None


def admin_required_response():
    current = current_user()
    if not current:
        return jsonify({'error': '推荐功能需要登录'}), 401
    if current.get('role') != 'admin':
        return jsonify({'error': '权限不足'}), 403
    return None


def current_user_id():
    return int((current_user() or {}).get('uid') or 0)


def ok(**payload):
    return jsonify({'success': True, **payload})


def is_dry_run():
    raw = request.args.get('dry_run')
    if raw is None:
        raw = json_data().get('dry_run')
    return str(raw).strip().lower() in ('1', 'true', 'yes', 'on')


def csv_rows(file_storage):
    text = file_storage.read().decode('utf-8-sig')
    return list(csv.DictReader(io.StringIO(text)))


def xlsx_rows(file_storage):
    from openpyxl import load_workbook

    workbook = load_workbook(file_storage, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        headers = [str(value or '').strip() for value in rows[0]] if rows else []
        return [dict(zip(headers, row)) for row in rows[1:] if any(row)]
    finally:
        workbook.close()


def attraction_csv(items):
    output = io.StringIO()
    fields = ['name', 'category', 'summary', 'tags', 'visit_minutes', 'difficulty', 'indoor', 'enabled']
    writer = csv.DictWriter(output, fieldnames=fields)
    writer.writeheader()
    for item in items:
        row = {key: item.get(key, '') for key in fields}
        row['tags'] = '|'.join(item.get('tags') or [])
        writer.writerow(row)
    return '\ufeff' + output.getvalue()


def attraction_xlsx(items):
    from openpyxl import Workbook

    workbook = Workbook()
    sheet = workbook.active
    fields = ['name', 'category', 'summary', 'tags', 'visit_minutes', 'difficulty', 'indoor', 'enabled']
    sheet.append(fields)
    for item in items:
        row = [('|'.join(item.get(key) or []) if key == 'tags' else item.get(key, '')) for key in fields]
        sheet.append(row)
    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return output
